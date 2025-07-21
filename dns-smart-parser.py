import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
import time
import random

# === 1. СБОР СПИСКА ССЫЛОК, НАЗВАНИЙ И ЦЕН ===

options = uc.ChromeOptions()
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--start-maximized")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

driver = uc.Chrome(options=options, version_main=138)

# Создание Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Products"
ws.append(["Название", "Цена (₽)", "Ссылка", "Процессор", "Встроенная память", "ОЗУ"])

# Начальная страница
base_url = "https://www.dns-shop.ru/catalog/17a8a01d16404e77/smartfony/"
driver.get("https://www.dns-shop.ru/")

# Установка куки
cookies = [{
    "name": "current_path",
    "value": "605bfdc517d7e9e23947448a9bf1ce16ac36b884434a3fdb10db053793c50392a%3A2%3A%7Bi%3A0%3Bs%3A12%3A%22current_path%22%3Bi%3A1%3Bs%3A115%3A%22%7B%22city%22%3A%2230b7c1f3-03fb-11dc-95ee-00151716f9f5%22%2C%22cityName%22%3A%22%5Cu041c%5Cu043e%5Cu0441%5Cu043a%5Cu0432%5Cu0430%22%2C%22method%22%3A%22manual%22%7D%22%3B%7D",
    "domain": ".dns-shop.ru",
    "path": "/",
    "expires": 1784575950
}]
for cookie in cookies:
    driver.add_cookie(cookie)

driver.get(base_url)
wait = WebDriverWait(driver, 30)

links = []

page = 1
while True:
    print(f"\n📄 Страница {page}")
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.catalog-products")))

    # Прокрутка
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(random.uniform(2, 3))
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    # Сбор карточек
    products = driver.find_elements(By.CSS_SELECTOR, "div.catalog-product")
    print(f"🔍 Найдено товаров: {len(products)}")

    for product in products:
        try:
            name_el = product.find_element(By.CSS_SELECTOR, "a.catalog-product__name")
            price_el = product.find_element(By.CSS_SELECTOR, "div.product-buy__price")

            name = name_el.text.strip()
            price_text = price_el.text.strip().replace(" ", "").replace("₽", "")
            price = int(price_text) if price_text.isdigit() else price_text
            product_url = name_el.get_attribute("href")

            if "/product/" in product_url:
                parts = product_url.split("/product/")[1].split("/")
                product_path = parts[0]
                product_slug = parts[1]
                characteristics_url = f"https://www.dns-shop.ru/product/characteristics/{product_path}/{product_slug}/"

                ws.append([name, price, characteristics_url, "", "", ""])
                links.append(characteristics_url)

        except Exception as e:
            print(f"⚠️ Ошибка при обработке товара: {e}")
            continue

    # Переход на следующую страницу
    try:
        next_btn = driver.find_element(By.CSS_SELECTOR, "a.pagination-widget__page-link_next")
        if "pagination-widget__page-link_disabled" in next_btn.get_attribute("class"):
            print("⛔ Последняя страница достигнута.")
            break
        driver.get(next_btn.get_attribute("href"))
        page += 1
        time.sleep(2)
    except:
        print("⛔ Кнопка 'Следующая' не найдена.")
        break

# Сохраняем предварительно
wb.save("products.xlsx")
print("\n✅ Первый этап завершён — данные сохранены")

# === 2. ПАРСИНГ ХАРАКТЕРИСТИК ===

target_keys = {
    "Модель процессора": "Процессор",
    "Объем встроенной памяти": "Встроенная память",
    "Объем оперативной памяти": "ОЗУ"
}

for i, url in enumerate(links, start=2):  # начинаем со 2-й строки
    print(f"\n🔎 Обработка характеристик: {url}")
    try:
        driver.get(url)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.product-characteristics-content")))

        # Клик по "Развернуть все"
        try:
            expand_btn = driver.find_element(By.CSS_SELECTOR, "button.product-characteristics__expand")
            driver.execute_script("arguments[0].click();", expand_btn)
            time.sleep(1)
        except:
            pass

        # Прокрутка
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)

        block = driver.find_element(By.CSS_SELECTOR, "div.product-characteristics-content")
        html = block.get_attribute("outerHTML")
        soup = BeautifulSoup(html, "html.parser")

        found_values = {}
        for item in soup.select("li.product-characteristics__spec"):
            name_el = item.select_one(".product-characteristics__spec-title")
            value_el = item.select_one(".product-characteristics__spec-value")
            if not name_el or not value_el:
                continue

            name = name_el.get_text(strip=True).replace(":", "")
            value = value_el.get_text(strip=True)

            if name in target_keys:
                found_values[target_keys[name]] = value

        # Запись в Excel
        ws.cell(row=i, column=4).value = found_values.get("Процессор", "")
        ws.cell(row=i, column=5).value = found_values.get("Встроенная память", "")
        ws.cell(row=i, column=6).value = found_values.get("ОЗУ", "")

        time.sleep(1)

    except Exception as e:
        print(f"⚠️ Ошибка при парсинге характеристик: {e}")
        continue

# Сохраняем финальный Excel
wb.save("products.xlsx")
print("\n✅ Все данные и характеристики сохранены в products.xlsx")

driver.quit()
